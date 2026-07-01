"""
data_loader.py — Karben4 QM Yield Tool · workbook -> batch objects (increment 3)
=================================================================================
Reads the brewery's two tracking workbooks into typed batch objects engine.py
can consume directly. No web/CSV path yet — Ekos has no confirmed self-serve API
(see "Brewery QM Responses — 2026-06-24"), so workbook import is the pragmatic
near-term path; a CSV loader can sit alongside this one if an Ekos export shows up.

Two sources, joined by Batch Number:
  - Lauter_Checks_2.xlsx  -> one tab per batch: grain bill (engine.GrainItem) +
                             lauter inputs (engine.LauterInputs).
  - Brewery_Yields.xlsx   -> one column per batch in 'Brewhouse' (engine.BrewhouseInputs)
                             and 'Cellar' (engine.CellarInputs); not every batch
                             has both, or a matching Lauter_Checks tab.

A Batch can be partial (e.g. cellar data but no detailed lauter check) — callers
should check which fields are populated (None where a source is missing) before
calling the corresponding engine function.
"""
import os
from dataclasses import dataclass, field
from typing import Optional
import openpyxl

from engine import GrainItem, LauterInputs, SugarAdditions, BrewhouseInputs, CellarInputs


@dataclass
class Batch:
    batch_number: float
    beer: Optional[str] = None
    brew_date: Optional[object] = None
    grains: list = field(default_factory=list)            # list[GrainItem], from Lauter_Checks
    lauter: Optional[LauterInputs] = None
    brewhouse: Optional[BrewhouseInputs] = None
    cellar: Optional[CellarInputs] = None


def _num(v):
    return isinstance(v, (int, float))


def _batch_id(v):
    """Normalize a Batch Number cell to a join key. Brewery_Yields' 'Brewhouse' tab stores
    it as text ('251027.02') while Cellar + Lauter_Checks store it as a float — without this,
    the two workbooks silently fail to join (string keys never match float keys)."""
    if v is None or v == "":
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def load_lauter_checks(path: str) -> dict:
    """{batch_number: Batch} with grains + lauter populated — one workbook tab per batch."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for tab in wb.sheetnames:
        if tab in ("Analysis", "Grain", "Background"):
            continue
        ws = wb[tab]
        batch_no = _batch_id(ws.cell(5, 2).value)
        if batch_no is None:
            continue
        if batch_no in out:
            print(f"  [data_loader] WARNING: duplicate batch number {batch_no} "
                  f"({out[batch_no].beer!r} and {tab!r} both claim it) — keeping {tab!r}")
        grains = []
        col = 2
        while True:
            name = ws.cell(33, col).value
            if name in (None, ""):
                break
            wt = ws.cell(82, col).value
            if _num(wt) and wt > 0:
                grains.append(GrainItem(
                    name=str(name), weight_lb=wt,
                    cgdb_yield_pct=ws.cell(83, col).value,
                    moisture_pct=ws.cell(84, col).value,
                    mill_yield_class=(ws.cell(86, col).value or "N"),
                ))
            col += 1
        lauter = LauterInputs(
            strike_water_temp_f=ws.cell(6, 2).value,
            strike_water_vol_gal=ws.cell(7, 2).value,
            lauter_runoff_vol_bbl=ws.cell(8, 2).value,
            lauter_runoff_extract_p=ws.cell(9, 2).value,
        )
        out[batch_no] = Batch(batch_number=batch_no, beer=tab, brew_date=ws.cell(4, 2).value,
                               grains=grains, lauter=lauter)
    return out


def load_brewery_yields(path: str) -> dict:
    """{batch_number: (BrewhouseInputs|None, CellarInputs|None, beer|None)}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    bw, ce = wb["Brewhouse"], wb["Cellar"]
    out = {}

    for col in range(2, bw.max_column + 1):
        batch_no = _batch_id(bw.cell(19, col).value)
        if batch_no is None:
            continue
        eob_p, hops_lb = bw.cell(30, col).value, bw.cell(21, col).value
        runoff_vol, runoff_p = bw.cell(28, col).value, bw.cell(29, col).value
        if not all(_num(x) for x in (eob_p, hops_lb, runoff_vol, runoff_p)):
            continue
        sugars = SugarAdditions(
            brewers_crystals_lb=bw.cell(22, col).value or 0, dme_lb=bw.cell(23, col).value or 0,
            dextrose_lb=bw.cell(24, col).value or 0, sucrose_lb=bw.cell(25, col).value or 0,
            lactose_lb=bw.cell(26, col).value or 0, maltodextrin_lb=bw.cell(27, col).value or 0,
        )
        bh = BrewhouseInputs(runoff_vol, runoff_p, eob_p, hops_lb, sugars)
        rec = out.setdefault(batch_no, [None, None, None])
        rec[0] = bh
        rec[2] = bw.cell(18, col).value

    for col in range(2, ce.max_column + 1):
        batch_no = _batch_id(ce.cell(16, col).value)
        if batch_no is None:
            continue
        fv, og_p, dry_hops = ce.cell(17, col).value, ce.cell(18, col).value, ce.cell(19, col).value
        if not all(_num(x) for x in (fv, og_p, dry_hops)) or fv == 0:
            continue
        co_actual, bt_override, pkg_vol = ce.cell(20, col).value, ce.cell(21, col).value, ce.cell(22, col).value
        cl = CellarInputs(fv, og_p, dry_hops,
                           centrifuge_vol_out_actual_bbl=co_actual if _num(co_actual) else None,
                           bt_volume_start_of_packaging_bbl=bt_override if _num(bt_override) else None,
                           packaged_vol_bbl=pkg_vol if _num(pkg_vol) else None)
        rec = out.setdefault(batch_no, [None, None, None])
        rec[1] = cl
        if rec[2] is None:
            rec[2] = ce.cell(15, col).value

    return out


def load_batches(lauter_path: str, yields_path: str) -> dict:
    """Joins both workbooks on Batch Number into a single {batch_number: Batch} map."""
    batches = load_lauter_checks(lauter_path)
    yields = load_brewery_yields(yields_path)
    for batch_no, (bh, cl, beer) in yields.items():
        b = batches.get(batch_no)
        if b is None:
            b = Batch(batch_number=batch_no, beer=beer)
            batches[batch_no] = b
        b.brewhouse = bh
        b.cellar = cl
        if b.beer is None:
            b.beer = beer
    return batches


if __name__ == "__main__":
    LAUTER = os.path.join(os.path.dirname(__file__), "..", "..", "Inputs", "Lauter_Checks_2.xlsx")
    YIELDS = os.path.join(os.path.dirname(__file__), "..", "..", "Inputs", "Brewery_Yields.xlsx")
    batches = load_batches(LAUTER, YIELDS)
    n_lauter = sum(1 for b in batches.values() if b.grains)
    n_bh = sum(1 for b in batches.values() if b.brewhouse)
    n_ce = sum(1 for b in batches.values() if b.cellar)
    n_full = sum(1 for b in batches.values() if b.grains and b.brewhouse and b.cellar)
    print(f"Loaded {len(batches)} batches: {n_lauter} with lauter detail, "
          f"{n_bh} with knockout, {n_ce} with centrifuge-out, {n_full} with all three.")

    from engine import lauter_metrics, knockout_metrics, cellar_metrics
    for b in batches.values():
        if b.grains and b.lauter:
            lauter_metrics(b.lauter, b.grains)
        if b.brewhouse:
            knockout_metrics(b.brewhouse)
        if b.cellar:
            cellar_metrics(b.cellar)
    print("Smoke test OK: every loaded batch ran through its engine function(s) without error.")
