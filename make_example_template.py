"""
make_example_template.py — generate blank, brewery-agnostic workbook templates
==============================================================================
Produces two upload-ready template workbooks under templates/ that match the exact
cell layout data_loader.py reads, so a NEW brewery can fill in its own numbers and
upload them — with all the headers/labels in place and NO Karben4 proprietary data.

Each template carries ONE clearly-labelled EXAMPLE batch of obviously-made-up round
numbers (an "Example Pale Ale"), so the format is self-documenting. Delete or
overwrite that example column/tab and add your own.

    python make_example_template.py     # writes templates/*.xlsx

The layouts are Karben4's original (cell-positional) format, kept so the same loader
serves both. If you'd rather not touch a spreadsheet at all, skip these entirely and
type batches into the app's **Add batch** tab — the tool computes everything itself.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT_DIR = os.path.join(os.path.dirname(__file__), "templates")

HDR = Font(bold=True, size=12)
LBL = Font(bold=True)
NOTE = Font(italic=True, color="666666")
EXAMPLE_FILL = PatternFill("solid", fgColor="FFF6D5")   # pale gold = "example, replace me"
HDR_FILL = PatternFill("solid", fgColor="161B23")
HDR_FONT = Font(bold=True, size=12, color="7BE72D")


def _label(ws, row, text, col=1):
    c = ws.cell(row, col, text)
    c.font = LBL


def _example(ws, row, col, value):
    c = ws.cell(row, col, value)
    c.fill = EXAMPLE_FILL


def build_lauter_template(path):
    """One tab per batch; tab name = beer name. col B = batch scalars + grain #1,
    cols C,D,... = additional grains. Matches load_lauter_checks() cell positions."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Example Pale Ale"           # <- tab name IS the beer name

    # Banner
    ws.cell(1, 1, "LAUTER CHECK — one tab per batch. Tab name = beer name. "
                   "Fill column B (and C, D… for extra grains). Yellow = example, replace it.")
    ws.cell(1, 1).font = HDR
    ws.cell(2, 1, "Add a new tab (copy this one) for each batch. Rows/columns must stay in these positions.")
    ws.cell(2, 1).font = NOTE

    # Batch scalars (col B / col 2)
    _label(ws, 4, "Brew date (YYYY-MM-DD)");        _example(ws, 4, 2, "2026-01-15")
    _label(ws, 5, "Batch number");                  _example(ws, 5, 2, 100001.01)
    _label(ws, 6, "Strike water temp (°F)");        _example(ws, 6, 2, 165)
    _label(ws, 7, "Strike water volume (gal)");     _example(ws, 7, 2, 300)
    _label(ws, 8, "Lauter runoff volume (bbl)");    _example(ws, 8, 2, 17.6)
    _label(ws, 9, "Lauter runoff extract (°P)");    _example(ws, 9, 2, 16.0)

    # Grain bill — labels in col A, grain #1 in col B, grain #2 in col C, ...
    _label(ws, 32, "GRAIN BILL  →  one grain per column, starting column B")
    _label(ws, 33, "Grain name")
    _label(ws, 82, "Weight (lb)")
    _label(ws, 83, "CGDB extract yield (%)")
    _label(ws, 84, "Moisture (%)")
    _label(ws, 86, "Mill yield class (N = normal)")

    # Example: two grains
    grains = [
        ("Base Malt (2-Row)", 500, 80.0, 4.0, "N"),
        ("Crystal 40",         50, 75.0, 4.0, "N"),
    ]
    for i, (name, wt, cgdb, moist, mill) in enumerate(grains):
        col = 2 + i
        _example(ws, 33, col, name)
        _example(ws, 82, col, wt)
        _example(ws, 83, col, cgdb)
        _example(ws, 84, col, moist)
        _example(ws, 86, col, mill)

    ws.column_dimensions["A"].width = 34
    for col in "BCDE":
        ws.column_dimensions[col].width = 18

    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(path)


def build_yields_template(path):
    """Two tabs: Brewhouse + Cellar. One COLUMN per batch (col B = batch #1,
    col C = batch #2, ...). Matches load_brewery_yields() cell positions."""
    wb = openpyxl.Workbook()

    bw = wb.active
    bw.title = "Brewhouse"
    bw.cell(1, 1, "BREWHOUSE (knockout) — one COLUMN per batch, starting column B. "
                   "Yellow = example, replace it.")
    bw.cell(1, 1).font = HDR
    bh_rows = [
        (18, "Beer name",                         "Example Pale Ale"),
        (19, "Batch number",                      100001.01),
        (21, "Hops (lb)",                          10),
        (22, "Brewers crystals (lb)",              0),
        (23, "DME (lb)",                           0),
        (24, "Dextrose (lb)",                      0),
        (25, "Sucrose (lb)",                       0),
        (26, "Lactose (lb)",                       0),
        (27, "Maltodextrin (lb)",                  0),
        (28, "Kettle runoff volume (bbl)",         17.6),
        (29, "Kettle runoff extract (°P)",         16.0),
        (30, "End-of-boil extract (°P)",           12.5),
    ]
    for r, label, ex in bh_rows:
        _label(bw, r, label)
        _example(bw, r, 2, ex)
    bw.column_dimensions["A"].width = 30
    for col in "BCDE":
        bw.column_dimensions[col].width = 16

    ce = wb.create_sheet("Cellar")
    ce.cell(1, 1, "CELLAR — one COLUMN per batch, starting column B. "
                   "Last three rows are optional. Yellow = example, replace it.")
    ce.cell(1, 1).font = HDR
    ce_rows = [
        (15, "Beer name",                              "Example Pale Ale"),
        (16, "Batch number",                           100001.01),
        (17, "Fermenter volume (bbl)",                 15.0),
        (18, "Original gravity (°P)",                  12.5),
        (19, "Dry hops (lb)",                          5),
        (20, "Centrifuge volume out (bbl) — optional", 14.0),
        (21, "BT volume at packaging (bbl) — optional",13.8),
        (22, "Packaged volume (bbl) — optional",       13.5),
    ]
    for r, label, ex in ce_rows:
        _label(ce, r, label)
        _example(ce, r, 2, ex)
    ce.column_dimensions["A"].width = 40
    for col in "BCDE":
        ce.column_dimensions[col].width = 16

    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(path)


def main():
    lauter = os.path.join(OUT_DIR, "Lauter_Checks_TEMPLATE.xlsx")
    yields = os.path.join(OUT_DIR, "Brewery_Yields_TEMPLATE.xlsx")
    build_lauter_template(lauter)
    build_yields_template(yields)
    print("Wrote:")
    print(" ", lauter)
    print(" ", yields)


if __name__ == "__main__":
    main()
