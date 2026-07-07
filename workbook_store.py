"""
workbook_store.py — durable .xlsx-backed batch store (auto-saves every batch)
=============================================================================
The tool's persistent "database" for a NEW brewery: a single human-readable
workbook (``brewery_data.xlsx``) that the app **creates on the first save and
rewrites on every add / edit / delete**, so data survives between sessions with
no manual export step. This replaces the old ``manual_batches.json`` (which is
migrated in automatically the first time this runs).

Two sheets, so a variable-length grain bill round-trips cleanly and a person can
still read/edit the file by hand:

  • "Batches" — one row per batch: scalars for the lauter, brewhouse (knockout)
                and cellar stages. A stage is considered *present* if any of its
                columns is filled; entirely-blank stage columns mean that stage
                wasn't recorded (round-trips None vs. present-with-zeros).
  • "Grains"  — one row per grain (batch_number + the five GrainItem fields),
                joined back to its batch on batch_number.

⚠️ Durability note: this persists on a real disk (a brewery laptop/server running
``streamlit run app.py``). On Streamlit Community Cloud the filesystem is ephemeral
— the file is wiped on restart/redeploy — so the hosted URL still needs a real
machine or a connected datastore for durable storage. See the README.
"""
import os
import json
import openpyxl

from data_loader import Batch
from engine import GrainItem, LauterInputs, SugarAdditions, BrewhouseInputs, CellarInputs

WORKBOOK_PATH = os.path.join(os.path.dirname(__file__), "brewery_data.xlsx")
LEGACY_JSON_PATH = os.path.join(os.path.dirname(__file__), "manual_batches.json")

# --- column layouts (header order == write/read order) -----------------------
LAUTER_COLS = ["strike_water_temp_f", "strike_water_vol_gal",
               "lauter_runoff_vol_bbl", "lauter_runoff_extract_p"]
# brewhouse: its own runoff pair (prefixed bh_ to disambiguate from lauter's),
# then eob + hops, then the six flattened sugar fields.
BREWHOUSE_COLS = ["bh_lauter_runoff_vol_bbl", "bh_lauter_runoff_extract_p",
                  "end_of_boil_extract_p", "kettle_whirlpool_hops_lb",
                  "brewers_crystals_lb", "dme_lb", "dextrose_lb",
                  "sucrose_lb", "lactose_lb", "maltodextrin_lb"]
CELLAR_COLS = ["effective_fv_wort_vol_bbl", "effective_fv_wort_og_p", "dry_hops_lb",
               "centrifuge_vol_out_actual_bbl", "bt_volume_start_of_packaging_bbl",
               "packaged_vol_bbl"]
BATCH_HEADERS = ["batch_number", "beer", "brew_date"] + LAUTER_COLS + BREWHOUSE_COLS + CELLAR_COLS
GRAIN_HEADERS = ["batch_number", "name", "weight_lb", "cgdb_yield_pct",
                 "moisture_pct", "mill_yield_class"]

SUGAR_FIELDS = ["brewers_crystals_lb", "dme_lb", "dextrose_lb",
                "sucrose_lb", "lactose_lb", "maltodextrin_lb"]


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _batch_id(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


# --- serialize a Batch -> flat row dicts -------------------------------------
def _batch_rows(b: Batch):
    """Return (batch_row: dict, grain_rows: list[dict]) for one Batch."""
    row = {h: None for h in BATCH_HEADERS}
    row["batch_number"] = b.batch_number
    row["beer"] = b.beer
    row["brew_date"] = None if b.brew_date is None else str(b.brew_date)

    if b.lauter is not None:
        for f in LAUTER_COLS:
            row[f] = getattr(b.lauter, f)
    if b.brewhouse is not None:
        row["bh_lauter_runoff_vol_bbl"] = b.brewhouse.lauter_runoff_vol_bbl
        row["bh_lauter_runoff_extract_p"] = b.brewhouse.lauter_runoff_extract_p
        row["end_of_boil_extract_p"] = b.brewhouse.end_of_boil_extract_p
        row["kettle_whirlpool_hops_lb"] = b.brewhouse.kettle_whirlpool_hops_lb
        for f in SUGAR_FIELDS:
            row[f] = getattr(b.brewhouse.sugars, f)
    if b.cellar is not None:
        for f in CELLAR_COLS:
            row[f] = getattr(b.cellar, f)

    grains = [{"batch_number": b.batch_number, "name": g.name, "weight_lb": g.weight_lb,
               "cgdb_yield_pct": g.cgdb_yield_pct, "moisture_pct": g.moisture_pct,
               "mill_yield_class": g.mill_yield_class} for g in (b.grains or [])]
    return row, grains


# --- deserialize flat rows -> Batch ------------------------------------------
def _row_to_batch(row: dict, grain_rows: list) -> Batch:
    def present(cols):
        return any(row.get(c) is not None for c in cols)

    lauter = None
    if present(LAUTER_COLS):
        lauter = LauterInputs(**{f: _num(row.get(f)) or 0.0 for f in LAUTER_COLS})

    brewhouse = None
    if present(BREWHOUSE_COLS):
        brewhouse = BrewhouseInputs(
            lauter_runoff_vol_bbl=_num(row.get("bh_lauter_runoff_vol_bbl")) or 0.0,
            lauter_runoff_extract_p=_num(row.get("bh_lauter_runoff_extract_p")) or 0.0,
            end_of_boil_extract_p=_num(row.get("end_of_boil_extract_p")) or 0.0,
            kettle_whirlpool_hops_lb=_num(row.get("kettle_whirlpool_hops_lb")) or 0.0,
            sugars=SugarAdditions(**{f: _num(row.get(f)) or 0.0 for f in SUGAR_FIELDS}),
        )

    cellar = None
    if present(CELLAR_COLS):
        cellar = CellarInputs(**{f: _num(row.get(f)) or 0.0 for f in CELLAR_COLS})

    grains = [GrainItem(name=g.get("name") or "", weight_lb=_num(g.get("weight_lb")) or 0.0,
                        cgdb_yield_pct=_num(g.get("cgdb_yield_pct")) or 0.0,
                        moisture_pct=_num(g.get("moisture_pct")) or 0.0,
                        mill_yield_class=(g.get("mill_yield_class") or "N"))
              for g in grain_rows]

    return Batch(batch_number=_batch_id(row.get("batch_number")),
                 beer=row.get("beer"), brew_date=row.get("brew_date"),
                 grains=grains, lauter=lauter, brewhouse=brewhouse, cellar=cellar)


# --- workbook <-> batches (shared by the local + SharePoint stores) ----------
def batches_to_workbook(batches: dict) -> "openpyxl.Workbook":
    """Build an in-memory workbook (Batches + Grains sheets) from {num: Batch}.
    Shared by _write_all (local disk) and sharepoint_store (Graph upload)."""
    wb = openpyxl.Workbook()
    bs = wb.active
    bs.title = "Batches"
    bs.append(BATCH_HEADERS)
    gs = wb.create_sheet("Grains")
    gs.append(GRAIN_HEADERS)
    for b in batches.values():
        row, grains = _batch_rows(b)
        bs.append([row[h] for h in BATCH_HEADERS])
        for g in grains:
            gs.append([g[h] for h in GRAIN_HEADERS])
    bs.freeze_panes = "A2"
    gs.freeze_panes = "A2"
    for ws in (bs, gs):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 32)
    return wb


def workbook_to_batches(wb: "openpyxl.Workbook") -> dict:
    """Parse an in-memory workbook (Batches + Grains sheets) -> {num: Batch}.
    Shared by load_batches (local disk) and sharepoint_store (Graph download)."""
    if "Batches" not in wb.sheetnames:
        return {}
    rows = list(wb["Batches"].iter_rows(values_only=True))
    if not rows:
        return {}
    headers = list(rows[0])
    batch_rows = [dict(zip(headers, r)) for r in rows[1:] if any(v is not None for v in r)]

    grain_by_batch = {}
    if "Grains" in wb.sheetnames:
        grows = list(wb["Grains"].iter_rows(values_only=True))
        if grows:
            gh = list(grows[0])
            for r in grows[1:]:
                gd = dict(zip(gh, r))
                bid = _batch_id(gd.get("batch_number"))
                if bid is not None:
                    grain_by_batch.setdefault(bid, []).append(gd)

    out = {}
    for row in batch_rows:
        bid = _batch_id(row.get("batch_number"))
        if bid is None:
            continue
        out[bid] = _row_to_batch(row, grain_by_batch.get(bid, []))
    return out


# --- local-disk read / write -------------------------------------------------
def _write_all(batches: dict, path: str = WORKBOOK_PATH):
    wb = batches_to_workbook(batches)
    tmp = path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, path)   # atomic-ish: never leave a half-written file


def load_batches(path: str = WORKBOOK_PATH) -> dict:
    """{batch_number: Batch} read back from the workbook. Empty if it doesn't exist
    (first run). Migrates a legacy manual_batches.json in on first use."""
    if not os.path.exists(path):
        migrated = _migrate_legacy_json(path)
        if not migrated:
            return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    return workbook_to_batches(wb)


def store_label() -> str:
    """Human label for the sidebar 'Your saved data' status."""
    return f"`{os.path.basename(WORKBOOK_PATH)}` (next to the app)"


def save_batch(batch: Batch, path: str = WORKBOOK_PATH):
    """Upsert one batch (keyed by batch_number) and rewrite the whole workbook."""
    batches = load_batches(path)
    batches[_batch_id(batch.batch_number)] = batch
    _write_all(batches, path)


def delete_batch(batch_number, path: str = WORKBOOK_PATH):
    batches = load_batches(path)
    batches.pop(_batch_id(batch_number), None)
    _write_all(batches, path)


def _migrate_legacy_json(path: str = WORKBOOK_PATH) -> bool:
    """One-time: if the workbook is absent but an old manual_batches.json exists,
    import those batches into the workbook. Returns True if anything was migrated."""
    if os.path.exists(path) or not os.path.exists(LEGACY_JSON_PATH):
        return False
    try:
        with open(LEGACY_JSON_PATH) as f:
            raw = json.load(f)
    except Exception:
        return False
    if not raw:
        return False
    batches = {}
    for num, d in raw.items():
        lauter = LauterInputs(**d["lauter"]) if d.get("lauter") else None
        brewhouse = None
        if d.get("brewhouse"):
            bh = dict(d["brewhouse"])
            sugars = SugarAdditions(**bh.pop("sugars")) if bh.get("sugars") else SugarAdditions(
                **{f: 0.0 for f in SUGAR_FIELDS})
            brewhouse = BrewhouseInputs(sugars=sugars, **bh)
        cellar = CellarInputs(**d["cellar"]) if d.get("cellar") else None
        grains = [GrainItem(**g) for g in d.get("grains", [])]
        batches[_batch_id(num)] = Batch(batch_number=_batch_id(num), beer=d.get("beer"),
                                        brew_date=d.get("brew_date"), grains=grains,
                                        lauter=lauter, brewhouse=brewhouse, cellar=cellar)
    _write_all(batches, path)
    return True
